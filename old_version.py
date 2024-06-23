import threading
import time
from functools import partial
import pyautogui as pg
from collections import defaultdict
import openpyxl
import datetime
from datetime import timedelta
import os
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkinter import filedialog as fd
import keyboard
from pprint import pprint

editing_state = False
END = False


class EditWindow(tk.Toplevel):
    def __init__(self, values=None):
        global editing_state
        super().__init__()
        self.values = values
        self.geometry("330x450")
        self.title('Редактировать')
        self.resizable(False, False)
        self.grab_set()
        self.geometry("+{}+{}".format(root.winfo_rootx() + 50, root.winfo_rooty() + 50))
        self.route_frame = ttk.Frame(self)
        self.route_label = ttk.Label(self.route_frame, text="Наряд:")
        self.route_entry = tk.Entry(self.route_frame, justify='left', width=35)
        self.direction_frame = ttk.Frame(self, padding=3)
        self.direction_label = tk.Label(self.direction_frame, text="Направление:")
        self.direction_entry = tk.Entry(self.direction_frame, width=35)
        self.plan_frame = ttk.Frame(self, padding=3)
        self.plan_label = tk.Label(self.plan_frame, text="План.время начала рейса:")
        self.plan_entry = tk.Entry(self.plan_frame, width=35)
        self.fact_frame = ttk.Frame(self, padding=3)
        self.fact_label = tk.Label(self.fact_frame, text="Факт.время начала рейса:")
        self.fact_entry = tk.Entry(self.fact_frame, width=35)
        self.bus_numb_frame = ttk.Frame(self, padding=3)
        self.bus_numb_label = tk.Label(self.bus_numb_frame, text="Гаражный номер:")
        self.bus_numb_entry = tk.Entry(self.bus_numb_frame, width=35)
        self.problem_frame = ttk.Frame(self, padding=3)
        self.problem_label = tk.Label(self.problem_frame, text="Проблема:")
        self.problem_entry = tk.Entry(self.problem_frame, width=35)
        self.screen_frame = ttk.Frame(self, padding=3)
        self.screen_label = tk.Label(self.screen_frame, text="Скрин:")
        self.screen_entry = tk.Entry(self.screen_frame, width=35)
        self.btn_frame = ttk.Frame(self, padding=3)
        self.btn_edit = ttk.Button(self.btn_frame, text='Редактировать', command=self.edit)
        self.fields = [self.route_entry, self.direction_entry, self.plan_entry, self.fact_entry,
                  self.bus_numb_entry, self.problem_entry, self.screen_entry]
        self.pack_items()
        self.fill_out_fields()
        editing_state = True
        res_panel.state = 0
        self.protocol('WM_DELETE_WINDOW', self.finish_editing)

    def finish_editing(self):
        global editing_state
        res_panel.state = 1
        editing_state = False
        self.destroy()

    def pack_items(self):
        self.route_frame.grid(row=0, column=0, padx=20)
        self.direction_frame.grid(row=1, column=0, padx=20)
        self.plan_frame.grid(row=2, column=0, padx=20)
        self.fact_frame.grid(row=3, column=0, padx=20)
        self.bus_numb_frame.grid(row=4, column=0, padx=20)
        self.problem_frame.grid(row=5, column=0, padx=20)
        self.screen_frame.grid(row=6, column=0, padx=20)
        self.btn_frame.grid(row=7, column=0, pady=20, sticky='nswe',)

        self.route_label.pack(anchor='w')
        self.route_entry.pack()
        self.direction_label.pack(anchor='w')
        self.direction_entry.pack()
        self.plan_label.pack(anchor='w')
        self.plan_entry.pack()
        self.fact_label.pack(anchor='w')
        self.fact_entry.pack()
        self.bus_numb_label.pack(anchor='w')
        self.bus_numb_entry.pack()
        self.problem_label.pack(anchor='w')
        self.problem_entry.pack()
        self.screen_label.pack(anchor='w')
        self.screen_entry.pack()
        self.btn_edit.pack(side=tk.RIGHT, padx=15)

        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        self.rowconfigure(2, weight=1)
        self.rowconfigure(3, weight=1)
        self.rowconfigure(4, weight=1)
        self.rowconfigure(5, weight=1)
        self.rowconfigure(6, weight=1)
        self.rowconfigure(7, weight=1)
        self.columnconfigure(0, weight=1)

    def fill_out_fields(self):
        for i, field in enumerate(self.fields):
            field.insert(0, str(self.values[i]))

    def get_data(self):
        res = []
        for field in self.fields:
            value = field.get()
            try:
                value = int(value) if '_' not in str(value) else str(value)
            except ValueError:
                pass
            res.append(value)
        return res

    @staticmethod
    def del_screen_or_not():
        screen_path = table.item(str(table.current_item))['values'][7]
        screen_name = screen_path.split('\\')[-1]
        res = messagebox.askyesno(title='Уведомление', message=f'Удалить скриншот "{screen_name}"?')
        if res:
            os.remove(screen_path)
            table.set(str(table.current_item), 7, '')

    def edit(self):
        item_id = str(table.current_item)
        old_values = table.item(str(table.current_item))['values']
        edited_values = self.get_data()
        if old_values[:7] != edited_values:
            for i, value in enumerate(self.get_data()):
                table.set(str(table.current_item), i, value)
            new_queue = edited_values[0]
            new_route = self.get_route(new_queue)
            table.set(str(table.current_item), 8, new_route)
            self.change_counter(old_values, edited_values)
            if item_id not in table.edited_items:
                table.edited_items[item_id] = {
                    'old_values': old_values,
                }
        self.finish_editing()

    def get_route(self, queue):
        if '_' in str(queue):
            route = queue.split('_')[0]
        else:
            route = queue
        try:
            return int(route)
        except ValueError:
            return route.strip()

    @staticmethod
    def change_counter(old_values, edited_values):
        if not old_values[:7][-1] and str(edited_values[-1]).strip():
            res_panel.add_flight()
        if old_values[:7][-1] and not str(edited_values[-1]).strip():
            res_panel.subtract_flight()


class Table(ttk.Treeview):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.heading('queue', text='Наряд', anchor='w', image=icon_route)
        self.heading('direction', text='Направление', anchor='w', image=icon_direction, )
        self.heading("plan", text="План", anchor='w', image=icon_plan)
        self.heading("fact", text="Факт", anchor='w', image=icon_fact)
        self.heading("bus_numb", text="ТС", anchor='w', image=icon_bus)
        self.heading("problem", text="Проблема", anchor='w', image=icon_problem)
        self.heading('screen', text="Скрин", anchor='w', image=icon_screen)
        self.tag_configure('bold', font=('Arial', 13, 'bold'))
        self.column("queue", width=90, stretch=True)
        self.column("direction", stretch=True, width=155)
        self.column("plan", stretch=True, width=80)
        self.column("fact", stretch=True, width=100)
        self.column("bus_numb", stretch=True, width=100)
        self.column("problem", stretch=True, width=350)
        self.column("screen", stretch=True, width=100)
        self.tag_configure('gray_colored', background='#D3D3D3')
        self.tag_configure('green_colored', background='#98FB98')
        self.tag_configure('white_colored', background='white')
        self.tag_configure('blue_colored', background='#00BFFF')
        self.table_size = 0
        self.current_item = 0
        self.edited_items = {}

    def cancel(self):
        item_id = str(self.current_item)
        values = self.item(str(self.current_item))['values']
        screen_path, screen = values[7], values[6]
        if item_id in self.edited_items:
            if screen_path:
                os.remove(screen_path)
            for i, value in enumerate(self.edited_items[item_id]['old_values']):
                self.set(item_id, i, value)
            if self.current_item + 1 < self.table_size:
                self.current_item += 1
                self.color(-1)
                self.check()
                self.yview_scroll(1, 'units')
            if screen:
                res_panel.subtract_flight()
            del self.edited_items[item_id]

    def item_selected(self, event):
        if len(self.selection()):
            offset = self.current_item - int(self.selection()[0])
            if offset:
                selected_item = self.current_item - offset
                self.selection_remove(str(selected_item))
                self.current_item = selected_item
                self.color(offset)
                self.check()
            else:
                self.selection_remove(str(self.current_item))

    def up(self):
        if self.current_item - 1 >= 0:
            exc = self.item(str(self.current_item))['values'][10]
            self.current_item -= 1
            self.color(1, exc)
            self.check()
            self.yview_scroll(-1, 'units')

    def take_action(self, action=None):
        if self.current_item > self.table_size - 1:
            return
        values = self.item(str(self.current_item))['values']
        if action == 'Есть' and values[6]:
            screen_path = self.make_screenshot(values, True)
            self.set(str(self.current_item), 6, action)
            self.set(str(self.current_item), 7, screen_path)
            self.set(str(self.current_item), 10, 1)
            self.down(True)
        elif action == 'Есть' and not values[6]:
            self.append_to_edited(values)
            screen_path = self.make_screenshot(values)
            self.set(str(self.current_item), 7, screen_path)
            self.set(str(self.current_item), 6, action)
            self.down()
        elif action == 'Видео' and not values[6]:
            self.append_to_edited(values)
            self.mark_as_video()
            self.set(str(self.current_item), 6, action)
            self.down()
        elif action == 'Видео' and values[6] == 'Есть':
            os.remove(self.item(str(self.current_item))['values'][7])
            self.set(str(self.current_item), 7, '')
            self.mark_as_video(True)
            self.set(str(self.current_item), 6, action)
            self.set(str(self.current_item), 10, 1)
            self.down(True)
        else:
            self.down()

    def append_to_edited(self, values):
        if str(self.current_item) not in self.edited_items:
            self.edited_items[str(self.current_item)] = {
                'old_values': values,
            }

    def check(self):
        screen_path = self.item(str(self.current_item))['values'][7]
        if screen_path:
            button7['state'] = 'normal'
        else:
            button7['state'] = 'disabled'

    def show_screen(self):
        screen_path = self.item(str(self.current_item))['values'][7]
        os.startfile(screen_path)

    def color_after_action(self, offset, exc):
        self.item(str(self.current_item), tags=('gray_colored',))
        if self.item(str(self.current_item + offset))['values'][6] and not exc:
            self.item(str(self.current_item + offset), tags=('green_colored',))
            self.set(str(self.current_item + offset), 13, 'green_colored')
        elif self.item(str(self.current_item + offset))['values'][6] and exc:
            self.item(str(self.current_item + offset), tags=('blue_colored',))
            self.set(str(self.current_item + offset), 13, 'blue_colored')

    def color(self, offset):
        colour = self.item(str(self.current_item + offset))['values'][13]
        self.item(str(self.current_item), tags=('gray_colored',))
        self.item(str(self.current_item + offset), tags=(colour,))

    def down(self, edited=False):
        if self.current_item + 1 < self.table_size:
            exc = self.item(str(self.current_item))['values'][10]
            self.current_item += 1
            self.color_after_action(-1, exc)
            self.check()
            self.yview_scroll(1, 'units')
        else:
            if edited:
                self.set(str(self.current_item), 13, 'blue_colored')
            else:
                self.set(str(self.current_item), 13, 'green_colored')

    def fill_out_table(self, rd):
        counter = -1
        for position, flight in rd.get_flight():
            counter += 1
            route_queue = f'{flight["route"]}{"__" + flight["queue"] if flight["queue"] else ""}'
            values = (str(route_queue), flight['direction'], flight['plan'],
                      flight['fact'], flight['bus_numb'], flight['problem'],
                      '', '', flight["route"], position, 0, flight['row_numb'], flight['date'], 'white_colored')
            self.insert('', 'end', values=values, iid=str(counter))
        self.table_size = counter + 1
        if self.table_size:
            self.item(str(self.current_item), tags=('gray_colored',))
        else:
            block_buttons(*buttons)

    @staticmethod
    def make_screenshot(values, exc=False):
        route = values[8]
        date = values[12]
        if date not in os.listdir('скрины'):
            os.mkdir(rf'скрины\{date}')
        if str(route) not in os.listdir(fr'скрины\{date}'):
            os.mkdir(rf'скрины\{date}\{route}')
        screnshot_name = f'{values[2].replace(":", "_")} {values[4]} {values[5]}'
        pg.screenshot(rf'скрины\{date}\{route}\{screnshot_name}.jpg')
        if not exc:
            res_panel.add_flight()
        return rf'скрины\{date}\{route}\{screnshot_name}.jpg'

    @staticmethod
    def mark_as_video(exc=False):
        if not exc:
            res_panel.add_flight()


class LoadWindow(tk.Toplevel):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.geometry("850x250")
        self.title('ScreenShooter')
        self.resizable(False, False)
        self.grab_set()
        self.report_name = None
        self.final_report_name = None
        self.btn_upload = ttk.Button(self, text='Загрузить', state='disabled', command=self.close)
        self.frame_1 = None
        self.frame_2 = None
        # self.frame_3 = ttk.Frame(self, padding=3)

    def select_file(self, report_type=None):
        if report_type == 'unaccounted_flights':
            self.report_name = fd.askopenfilename()
            self.text_1.insert(1.0, self.report_name)
            self.check()
        else:
            self.final_report_name = fd.askopenfilename()
            self.text_2.insert(1.0, self.final_report_name)
            self.check()

    def check(self):
        if self.final_report_name and self.report_name:
            self.btn_upload['state'] = 'normal'

    def close(self):
        self.del_first_step()
        self.app.rd.file_path = self.report_name
        self.app.rd.final_report_path = self.final_report_name
        Thread(target=self.app.run_reader).start()

    def set(self):
        self.frame_1 = ttk.Frame(self, padding=3)
        self.frame_2 = ttk.Frame(self, padding=3)
        btn_1 = ttk.Button(self.frame_1, text='Найти', command=lambda: self.select_file('unaccounted_flights'))
        btn_2 = ttk.Button(self.frame_2, text='Найти', command=lambda: self.select_file())
        label_1 = ttk.Label(self.frame_1, text='Файл с неучтёнными рейсами:', justify='left')
        label_2 = ttk.Label(self.frame_2, text='Файл со всеми рейсами:', justify='left')
        self.text_1 = tk.Text(self.frame_1, height=1, width=65, background='white', font=('Arial', 11))
        self.text_2 = tk.Text(self.frame_2, height=1, width=65, background='white', font=('Arial', 11))
        label_1.grid(row=0, column=0, columnspan=2, sticky='we')
        self.text_1.grid(row=1, column=0)
        btn_1.grid(row=1, column=1)
        label_2.grid(row=0, column=0, columnspan=2, sticky='we')
        self.text_2.grid(row=1, column=0)
        btn_2.grid(row=1, column=4)
        self.frame_1.pack()
        self.frame_2.pack()
        self.btn_upload.pack(side='right', padx=10)


    def end(self):
        self.destroy()
        self.app.table.fill_out_table(self.app.rd)


class ResultPanel:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.main_frame = ttk.Frame(root, relief='groove')
        self.progress_var = tk.IntVar()
        self.progressbar = ttk.Progressbar(orient="horizontal", length=700, variable=self.progress_var)
        self.speed_counter = tk.StringVar(value='0')
        self.flights_counter = tk.IntVar()
        self.completed_counter = tk.IntVar()
        self.remaining_counter = tk.IntVar()
        self.worked_hours_counter = tk.StringVar(value='00:00:00')
        self.predict_counter = tk.StringVar(value='00:00:00')
        self.time_counter = 0
        self.state = 0

    def prepare_panel(self):
        speed_frame = ttk.Frame(self.main_frame, relief='groove', borderwidth=1, padding=3)
        predict_frame = ttk.Frame(self.main_frame, relief='groove', borderwidth=1, padding=3)
        flights_frame = ttk.Frame(self.main_frame, relief='groove', borderwidth=1, padding=3)
        completed_frame = ttk.Frame(self.main_frame, relief='groove', borderwidth=1, padding=3)
        remaining_frame = ttk.Frame(self.main_frame, relief='groove', borderwidth=1, padding=3)
        worked_hours_frame = ttk.Frame(self.main_frame, relief='groove', borderwidth=1, padding=3)

        speed_label = ttk.Label(speed_frame, text="Cкорость(cкр/мин):", font=("Arial", 12))
        predict_label = ttk.Label(predict_frame, text="Прогноз:", font=("Arial", 12))
        flights_label = ttk.Label(flights_frame, text="Всего:", font=("Arial", 12))
        completed_label = ttk.Label(completed_frame, text="Сделано:", font=("Arial", 12))
        remaining_label = ttk.Label(remaining_frame, text="Осталось:", font=("Arial", 12))
        worked_hours_label = ttk.Label(worked_hours_frame, text="Отработано:", font=("Arial", 12))

        speed_label.pack(side=tk.LEFT)
        predict_label.pack(side=tk.LEFT)
        flights_label.pack(side=tk.LEFT)
        completed_label.pack(side=tk.LEFT)
        remaining_label.pack(side=tk.LEFT)
        remaining_label.pack(side=tk.LEFT)
        worked_hours_label.pack(side=tk.LEFT)

        speed_display = ttk.Label(speed_frame, textvariable=self.speed_counter, font=("Arial", 18))
        flights_display = ttk.Label(flights_frame, textvariable=self.flights_counter, font=("Arial", 18))
        completed_display = ttk.Label(completed_frame, textvariable=self.completed_counter, font=("Arial", 18))
        remaining_display = ttk.Label(remaining_frame, textvariable=self.remaining_counter, font=("Arial", 18))
        worked_hours_display = ttk.Label(worked_hours_frame, textvariable=self.worked_hours_counter, font=("Arial", 18))
        predict_display = ttk.Label(predict_frame, textvariable=self.predict_counter, font=("Arial", 18))

        speed_display.pack(side=tk.RIGHT)
        flights_display.pack(side=tk.RIGHT)
        completed_display.pack(side=tk.RIGHT)
        remaining_display.pack(side=tk.RIGHT)
        predict_display.pack(side=tk.RIGHT)
        worked_hours_display.pack(side=tk.RIGHT)

        speed_frame.grid(row=0, column=0, padx=3, sticky='NSEW')
        worked_hours_frame.grid(row=0, column=1, padx=3, sticky='NSEW')
        predict_frame.grid(row=0, column=2, padx=3, sticky='NSEW')
        flights_frame.grid(row=0, column=3, padx=3, sticky='NSEW')
        completed_frame.grid(row=0, column=4, padx=3, sticky='NSEW')
        remaining_frame.grid(row=0, column=5, padx=3, sticky='NSEW')

        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.columnconfigure(1, weight=1)
        self.main_frame.columnconfigure(2, weight=1)
        self.main_frame.columnconfigure(3, weight=1)
        self.main_frame.columnconfigure(4, weight=1)
        self.main_frame.columnconfigure(5, weight=1)

        self.main_frame.grid(row=0, column=0, sticky='nsew')
        self.main_frame.rowconfigure(1, weight=1)

    def add_flight(self):
        self.completed_counter.set(self.completed_counter.get() + 1)
        self.remaining_counter.set(self.flights_counter.get() - self.completed_counter.get())
        speed = self.calc_speed()
        predict_value = 60 * self.remaining_counter.get() / speed
        self.predict_counter.set(str(timedelta(seconds=int(predict_value))))
        progress_value = 100 * self.completed_counter.get() / self.flights_counter.get()
        self.progress_var.set(int(progress_value))
        if not self.remaining_counter.get():
            show_inf()

    def subtract_flight(self):
        self.completed_counter.set(self.completed_counter.get() - 1)
        self.remaining_counter.set(self.flights_counter.get() - self.completed_counter.get())
        speed = self.speed_counter.get()
        predict_value = 60 * self.remaining_counter.get() / float(speed)
        self.predict_counter.set(str(timedelta(seconds=int(predict_value))))
        progress_value = 100 * self.completed_counter.get() / self.flights_counter.get()
        self.progress_var.set(int(progress_value))

    def run_time(self):
        self.time_counter += 1
        self.worked_hours_counter.set(str(timedelta(seconds=self.time_counter)))
        if self.state:
            self.root.after(1000, self.run_time)
        if editing_state:
            self.root.after(1000, self.run_time)

    def start(self):
        block_buttons(button1)
        activate_buttons(*buttons[1:])
        self.state = 1
        self.run_time()

    def pause(self):
        block_buttons(*buttons[1:])
        activate_buttons(button1)
        self.state = 0

    def calc_speed(self):
        try:
            past_minutes = self.time_counter / 60
            speed = round(self.completed_counter.get() / past_minutes, 1)
            self.speed_counter.set(str(speed))
            return speed
        except ZeroDivisionError:
            return 100.0


def show_error(message):
    messagebox.showerror('Ошибка', message)


def show_inf():
    messagebox.showinfo('Уведомление', 'Ура!Разбор рейсов закончен!')


def open_warning():
    messagebox.showwarning(title="Предупреждение", message="У вас открыт файл эксель с неучтенными рейсами."
                                                           "Для корректного завершения программы закройте его и нажмите 'Ок'")


def screen_by_space(event):
    if res_panel.state:
        table.take_action('Есть')


def get_value(val):
    if type(val) == datetime.time:
        return f"{val.hour:>02}:{val.minute:>02}"
    return val


def find_report():
    for file in os.listdir():
        if file.endswith('.xlsx'):
            return file
    raise FileNotFoundError


class Reader:
    def __init__(self, file_path=None):
        self.file_path = file_path
        self.final_report_path = None
        self.dict_problems = defaultdict(lambda: defaultdict(list))
        self.wb = None
        self.date = None
        self.total = 0

    def read(self, load_window: LoadWindow):
        self.wb = openpyxl.load_workbook(self.file_path)
        sheet = self.wb.active
        counter = 0
        for row in sheet.rows:
            counter += 1
            if row[0].value != 'Дата' and str(row[7].value).strip() in ('ОБРЫВ БЛОКА ГЛОНАСС', 'ПРОБКИ', 'РЕЙС ВЫПОЛНЕН '
                                                                                            'ПРАВИЛЬНО') \
                    and not row[8].value:
                if self.date is None and row[0].value:
                    self.convert_str_to_date(row[0].value)
                route = row[1].value
                if row[5].value:
                    self.total += 1
                    bus_numb = row[5].value
                    self.dict_problems[route][bus_numb].append(
                        {
                            'date': f'{row[0].value.day:>02}.{row[0].value.month:>02}.{row[0].value.year}'
                            if type(row[0].value) == datetime.datetime else row[0].value,
                            'direction': row[2].value,
                            'plan': self.convert_str_to_time(row[3].value),
                            'fact': self.convert_str_to_time(row[4].value),
                            'problem': row[7].value,
                            'screen': None,
                            'row_numb': counter,
                            'queue': "",
                            'bus_numb': row[5].value,
                            'route': row[1].value,
                        }
                    )

        for route in self.dict_problems:
            for lst_flights in self.dict_problems[route].values():
                lst_flights.sort(key=lambda item: self.get_seconds(item['plan']))
        load_window.progress_text_var.set(f'Файл {self.file_path} прочитан')
        load_window.progress_var.set(10)
        # pprint(self.dict_problems)

    @staticmethod
    def get_seconds(st):
        if st:
            hours, minutes = map(int, st.split(':'))
            res = hours * 60 + minutes
            return res
        return 0

    def convert_str_to_time(self, st: datetime.time):
        if st is None or st == '':
            return ''
        if type(st) == datetime.time or type(st) == datetime.datetime:
            return st.strftime('%H:%M')
        try:
            return datetime.time(*[int(val.strip()) for val in st.split(':')]).strftime('%H:%M')
        except Exception:
            return ''

    def convert_str_to_date(self, st):
        if type(st) == datetime.datetime:
            self.date = st.strftime('%d.%m.%Y')
        elif st.strip() and '.' in st:
            self.date = st.strip()

    def read_final_reports(self, load_window: LoadWindow):
        load_window.progress_text_var.set('Поиск нарядов....')
        handler = partial(update_progress_bar, load_window=load_window, q=q)
        root.bind('<<Updated>>', handler)
        wb = openpyxl.load_workbook(self.final_report_path)
        sheet = wb.active
        cur_bus = None
        search_res = False
        max_row = sheet.max_row
        step = max_row // 20
        step_counter = step
        for row in sheet.values:
            if step_counter == 0:
                root.event_generate('<<Updated>>', when='tail')
                time.sleep(0.5)

                step_counter = step
            if type(row[0]) == datetime.datetime and row[0].strftime('%d.%m.%Y') == self.date:
                if not search_res:
                    search_res = True
                route = self.convert_route_into_numb(row[2].replace('-КРС', ''))
                bus_numb = row[1]
                queue_ = str(row[10])
                if route in self.dict_problems and bus_numb in self.dict_problems[route] and bus_numb != cur_bus:
                    for flight in self.dict_problems[route][bus_numb]:
                        flight['queue'] = queue_

                    cur_bus = bus_numb
            if search_res and row[0].strftime('%d.%m.%Y') != self.date:
                break


        if load_window.progress_var.get() < 100:
            load_window.progress_var.set(100)
        load_window.progress_text_var.set('Наряды найдены')
        load_window.ok_btn['state'] = 'normal'

    def convert_route_into_numb(self, st):
        try:
            return int(st)
        except Exception:
            return st

    def get_flight(self):
        for route in self.dict_problems:
            for bus_numb in self.dict_problems[route]:
                for position, dict in enumerate(self.dict_problems[route][bus_numb]):
                    yield position, dict


class Writer:
    def __init__(self, reader: Reader):
        self.reader = reader

    def write(self):
        for i in range(table.table_size):
            item_id = str(i)
            if item_id in table.edited_items:
                values = table.item(item_id)['values']
                row_numb = values[11]
                self.reader.wb.active[f'B{row_numb}'] = values[8]
                self.reader.wb.active[f'C{row_numb}'] = values[1]
                self.reader.wb.active[f'D{row_numb}'] = values[2]
                self.reader.wb.active[f'E{row_numb}'] = values[3]
                self.reader.wb.active[f'F{row_numb}'] = values[4]
                self.reader.wb.active[f'H{row_numb}'] = values[5]
                self.reader.wb.active[f'I{row_numb}'] = values[6]
        self.reader.wb.save(self.reader.file_path)


def activate_buttons(*btns):
    for btn in btns:
        btn['state'] = 'normal'


def block_buttons(*btns):
    for btn in btns:
        btn['state'] = 'disabled'


def update_progress_bar(event, load_window, q=None):
    cur_value = load_window.progress_var.get()
    load_window.progress_var.set(cur_value + 5)

load_window = None
flight = None
root = tk.Tk()
res_panel = ResultPanel(root)
res_panel.prepare_panel()
res_panel.main_frame.grid(row=0, column=0, columnspan=6, sticky='nsew')
res_panel.progressbar.grid(row=1, column=0, columnspan=6, sticky='ew', pady=10)
button_frame = ttk.Frame(root)
icon_route = tk.PhotoImage(file=r'icons/icons8-автобусный-маршрут-20.png')
icon_direction = tk.PhotoImage(file=r'icons/icons8-направление-22.png')
icon_plan = tk.PhotoImage(file=r'icons/icons8-расписание-22.png')
icon_fact = tk.PhotoImage(file=r'icons/free-icon-shrug-5894030.png')
icon_bus = tk.PhotoImage(file=r'icons/icons8-автобус-22.png')
icon_problem = tk.PhotoImage(file=r'icons/icons8-внимание-22.png')
icon_screen = tk.PhotoImage(file=r'icons/icons8-скриншот-22.png')
video_icon = tk.PhotoImage(file=r'icons\icons8-видеозвонок-64.png')
screen_icon = tk.PhotoImage(file=r'icons\icons8-камера-50.png')
pass_step_icon = tk.PhotoImage(file=r'icons\icons8-пропустить-шаг-64.png')
cancel_icon = tk.PhotoImage(file=r'icons\icons8-отмена-50.png')
edit_icon = tk.PhotoImage(file=r'icons\icons8-редактировать-50.png')
show_icon = tk.PhotoImage(file=r'icons\icons8-глаз-50.png')
table_style = ttk.Style()
table_style.configure('Treeview', font=('Arial', 13), rowheight=60, separator=100)
heading_style = ttk.Style()
heading_style.configure('Treeview.Heading', font=('Arial', 12))

button_style = ttk.Style()
button_style.configure("mystyle.TButton",
                       font='Arial 13',
                       padding=10,

                       )

columns = ("queue", "direction", "plan", 'fact', 'bus_numb',
           'problem', 'screen', 'screen_path', 'route',
           'position', 'edited', 'row_id', 'date', 'colour')
table = Table(column=columns, show='headings', padding=10,
              displaycolumns=("queue", "direction", "plan", 'fact', 'bus_numb', 'problem', 'screen'))
table.bind('<<TreeviewSelect>>', table.item_selected)
button1 = ttk.Button(button_frame, text="Работать", state='disabled', command=res_panel.start, takefocus=False,
                     style='mystyle.TButton')
button2 = ttk.Button(button_frame, text="Перерыв", state='disabled', command=res_panel.pause, takefocus=False,
                     style='mystyle.TButton')
button3 = ttk.Button(button_frame, text="Скрин", command=lambda: table.take_action('Есть'),
                     image=screen_icon, compound='right', state='disabled', takefocus=False, style='mystyle.TButton')
button4 = ttk.Button(button_frame, text="Видео", command=lambda: table.take_action('Видео'), image=video_icon,
                     compound='right', state='disabled', takefocus=False, style='mystyle.TButton')

button5 = ttk.Button(button_frame, image=cancel_icon,
                     compound='image', takefocus=False, command=table.cancel,
                                                                        state='disabled')
button6 = ttk.Button(button_frame, image=edit_icon,
                     compound='image', takefocus=False, command=lambda: EditWindow(table.item(str(table.current_item))['values']), state='disabled')
button7 = ttk.Button(button_frame, image=show_icon,
                     compound='image', takefocus=False, command=table.show_screen, state='disabled')
buttons = [button1, button2, button3, button4, button5, button6]
button1.grid(row=1, column=0, sticky='nsew', pady=10, padx=10)
button2.grid(row=1, column=1, sticky='nsew', pady=10, padx=10)
button4.grid(row=2, column=0, columnspan=2, sticky='nsew', pady=5, padx=5)
button3.grid(row=3, column=0, columnspan=2, sticky='nsew', pady=10, padx=5)
button6.grid(row=3, column=4, pady=10, padx=20)
button5.grid(row=2, column=4, pady=10, padx=20)
button7.grid(row=1, column=4, pady=10, padx=20)
# button4.grid(row=6, column=0, columnspan=4, sticky='nsew', pady=10)
button_frame.grid(row=2, column=1, sticky='nsew', pady=130)
# text = tk.Text(root, font=("Arial", 11, 'roman'))
table.grid(row=2, column=2, columnspan=4, sticky='NSEW')
table.rowconfigure(0, pad=15)
scroll = tk.Scrollbar(command=table.yview)
scroll.grid(row=2, column=6, sticky='ns')
table.config(yscrollcommand=scroll.set)
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_columnconfigure(2, weight=1)
root.title('ScreenShooter')


def run():
    try:
        rd.read(load_window)
        rd.read_final_reports(load_window)
        activate_buttons(button1)
        res_panel.flights_counter.set(rd.total)
        res_panel.remaining_counter.set(rd.total)
        keyboard.on_press_key('space', screen_by_space)
    except PermissionError:
        if rd.wb:
            while True:
                try:
                    open_warning()
                    wr = Writer(rd)
                    wr.write()
                    break
                except PermissionError:
                    continue
        else:
            show_error("Открыт файл эксель с неучтенными рейсами.Закройте файл и перезапустите приложение!")


def new_window():
    global load_window
    load_window = LoadWindow(root)
    load_window.set()

def finish():
    try:
        wr = Writer(rd)
        wr.write()
    except PermissionError:
        if rd.wb:
            while True:
                try:
                    open_warning()
                    wr = Writer(rd)
                    wr.write()
                    break
                except PermissionError:
                    continue
        else:
            show_error("Открыт файл эксель с неучтенными рейсами.Закройте файл и перезапустите приложение!")
    finally:
        root.destroy()

try:
    os.mkdir('скрины') if 'скрины' not in os.listdir() else None
    rd = Reader()
    thread = threading.Thread(target=run)
    root.after(1000, new_window)
    root.protocol('WM_DELETE_WINDOW', finish)
    root.mainloop()
except FileNotFoundError:
    show_error("Не найден файл эксель с неучтенными рейсами.Загрузите файл и перезапустите приложение!")
except PermissionError:
    pass
except AttributeError:
    pass